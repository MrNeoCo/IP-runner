import subprocess
import time
from openpyxl import load_workbook
import os

# --- Load IPs from Excel file ---
excel_file = "ips.xlsx"

if not os.path.exists(excel_file):
    print(f"Excel file '{excel_file}' not found. Put it in the same folder.")
    exit()

wb = load_workbook(excel_file)
sheet = wb.active

ips = []

for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
    ip = row[0]
    if ip:
        ips.append(str(ip).strip())

if not ips:
    print("No IPs found in Excel. Add IPs in column A.")
    exit()

print(f"Found {len(ips)} IPs:", ips)

# --- Ping each IP in a separate CMD window ---
for ip in ips:
    print(f"Launching ping window for {ip}...")
    subprocess.Popen(["cmd", "/c", f"start cmd /k ping {ip} -t"])
    time.sleep(0.3)  # prevents window overlap

print("All ping windows started.")
